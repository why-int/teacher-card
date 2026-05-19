from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

PLAN_COLUMNS: list[tuple[str, str]] = [
    ("subject", "Предмет"),
    ("group", "Группа"),
    ("budget", "Тип (бюджет/вб)"),
    ("total_hours", "Всего часов за год"),
    ("sem1_hours", "1 семестр (часы)"),
    ("sem2_hours", "2 семестр (часы)"),
    ("sep", "Сентябрь"),
    ("oct_", "Октябрь"),
    ("nov", "Ноябрь"),
    ("dec", "Декабрь"),
    ("jan", "Январь"),
    ("feb", "Февраль"),
    ("mar", "Март"),
    ("apr", "Апрель"),
    ("may", "Май"),
    ("jun", "Июнь"),
]


@dataclass(frozen=True)
class RupLayout:
    subject: int = 0
    group: int = 1
    budget: int = 2
    sem1_hours: int = 7
    sem2_hours: int = 19
    total_hours: int = 28
    teacher: int = 30
    payment_type: int = 31
    sep: int = 32
    oct_: int = 33
    nov: int = 34
    dec: int = 35
    jan: int = 37
    feb: int = 38
    mar: int = 39
    apr: int = 40
    may: int = 41
    jun: int = 42


RUP = RupLayout()
TABLE_COLUMNS = [column for column, _ in PLAN_COLUMNS]


@dataclass
class PlanRecord:
    subject: str
    group: str
    budget: str
    teacher: str
    total_hours: float | int | None
    sem1_hours: float | int | None
    sem2_hours: float | int | None
    sep: float | int | None
    oct_: float | int | None
    nov: float | int | None
    dec: float | int | None
    jan: float | int | None
    feb: float | int | None
    mar: float | int | None
    apr: float | int | None
    may: float | int | None
    jun: float | int | None


def _extract_year_range(file_name: str) -> tuple[str, str]:
    match = re.search(r"(20\d{2})\s*-\s*(20\d{2})", file_name)
    if not match:
        raise ValueError(
            "Не удалось определить учебный год из названия файла РУП. "
            "Ожидается формат вида 2024-2025."
        )
    return match.group(1), match.group(2)


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).strip()


def _as_number_or_none(value: Any) -> float | int | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, str):
        text = value.strip().replace(",", ".")
        if not text:
            return None
        try:
            number = float(text)
        except ValueError:
            return None
    elif isinstance(value, (int, float)):
        number = float(value)
    else:
        return None

    if float(number).is_integer():
        return int(number)
    return float(number)


def _is_data_row(row: pd.Series) -> bool:
    subject = _clean_text(row.iloc[RUP.subject])
    group = _clean_text(row.iloc[RUP.group])
    teacher = _clean_text(row.iloc[RUP.teacher])
    payment_type = _clean_text(row.iloc[RUP.payment_type]).lower()
    if not subject or not group or not teacher:
        return False
    if subject.lower().startswith("итого"):
        return False
    # Строки с почасовой оплатой не включаем в карточки.
    if "почас" in payment_type:
        return False
    return True


def _sheet_name_from_teacher(teacher_name: str) -> str:
    text = re.sub(r"\s+", " ", teacher_name.strip())
    parts = [part for part in text.split(" ") if part]
    if not parts:
        return "Преподаватель"

    surname = parts[0]
    initials = ""
    if len(parts) >= 2:
        initials += parts[1][0].upper() + "."
    if len(parts) >= 3:
        initials += parts[2][0].upper() + "."

    candidate = f"{surname} {initials}".strip()
    invalid_chars = set("\\/*?:[]")
    candidate = "".join(ch for ch in candidate if ch not in invalid_chars)
    candidate = candidate.strip()
    if not candidate:
        candidate = "Преподаватель"
    return candidate[:31]


def _records_to_dataframe(records: list[PlanRecord]) -> pd.DataFrame:
    data = [
        {
            "subject": record.subject,
            "group": record.group,
            "budget": record.budget,
            "total_hours": record.total_hours,
            "sem1_hours": record.sem1_hours,
            "sem2_hours": record.sem2_hours,
            "sep": record.sep,
            "oct_": record.oct_,
            "nov": record.nov,
            "dec": record.dec,
            "jan": record.jan,
            "feb": record.feb,
            "mar": record.mar,
            "apr": record.apr,
            "may": record.may,
            "jun": record.jun,
        }
        for record in records
    ]
    return pd.DataFrame(data, columns=TABLE_COLUMNS)


def _record_from_row(row: pd.Series) -> PlanRecord:
    return PlanRecord(
        subject=_clean_text(row.iloc[RUP.subject]),
        group=_clean_text(row.iloc[RUP.group]),
        budget=_clean_text(row.iloc[RUP.budget]).lower(),
        teacher=_clean_text(row.iloc[RUP.teacher]),
        total_hours=_as_number_or_none(row.iloc[RUP.total_hours]),
        sem1_hours=_as_number_or_none(row.iloc[RUP.sem1_hours]),
        sem2_hours=_as_number_or_none(row.iloc[RUP.sem2_hours]),
        sep=_as_number_or_none(row.iloc[RUP.sep]),
        oct_=_as_number_or_none(row.iloc[RUP.oct_]),
        nov=_as_number_or_none(row.iloc[RUP.nov]),
        dec=_as_number_or_none(row.iloc[RUP.dec]),
        jan=_as_number_or_none(row.iloc[RUP.jan]),
        feb=_as_number_or_none(row.iloc[RUP.feb]),
        mar=_as_number_or_none(row.iloc[RUP.mar]),
        apr=_as_number_or_none(row.iloc[RUP.apr]),
        may=_as_number_or_none(row.iloc[RUP.may]),
        jun=_as_number_or_none(row.iloc[RUP.jun]),
    )


def _load_records_from_rup(rup_path: Path) -> list[PlanRecord]:
    raw_df = pd.read_excel(rup_path, sheet_name=0, header=None)
    records: list[PlanRecord] = []
    for _, row in raw_df.iterrows():
        if _is_data_row(row):
            records.append(_record_from_row(row))
    return records


def load_teacher_tables_from_rup(
    rup_file: str | Path,
) -> tuple[dict[str, pd.DataFrame], tuple[str, str]]:
    rup_path = Path(rup_file)
    if not rup_path.exists():
        raise FileNotFoundError(f"Файл РУП не найден: {rup_path}")

    year_start, year_end = _extract_year_range(rup_path.name)
    records = _load_records_from_rup(rup_path)
    if not records:
        raise ValueError("В РУП не найдены строки с данными для преподавателей.")

    by_teacher: dict[str, list[PlanRecord]] = {}
    for record in records:
        by_teacher.setdefault(record.teacher, []).append(record)

    tables: dict[str, pd.DataFrame] = {}
    for teacher, teacher_records in by_teacher.items():
        tables[teacher] = _records_to_dataframe(teacher_records)

    return tables, (year_start, year_end)


def _find_label_row(ws: Worksheet, label: str, column: int = 1) -> int:
    for row_idx in range(1, ws.max_row + 1):
        value = ws.cell(row_idx, column).value
        if value is None:
            continue
        if str(value).strip().lower().startswith(label.lower()):
            return row_idx
    raise ValueError(f"Не найдена строка с меткой '{label}' в листе '{ws.title}'.")


def _sum_values(values: list[float | int | None]) -> float:
    total = 0.0
    for value in values:
        if value is None:
            continue
        total += float(value)
    return total


def _copy_row_style(ws: Worksheet, src_row: int, dst_row: int, max_col: int) -> None:
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        dst._style = src._style  # noqa: SLF001
        if src.has_style:
            dst.number_format = src.number_format
            dst.font = src.font.copy()
            dst.fill = src.fill.copy()
            dst.border = src.border.copy()
            dst.alignment = src.alignment.copy()
            dst.protection = src.protection.copy()


def _write_record_row(ws: Worksheet, row_idx: int, record: PlanRecord) -> None:
    ws.cell(row_idx, 1).value = record.subject
    ws.cell(row_idx, 2).value = record.group
    ws.cell(row_idx, 3).value = record.total_hours
    ws.cell(row_idx, 11).value = record.total_hours
    ws.cell(row_idx, 12).value = record.sem1_hours
    ws.cell(row_idx, 14).value = record.sem2_hours
    ws.cell(row_idx, 16).value = record.sep
    ws.cell(row_idx, 17).value = record.oct_
    ws.cell(row_idx, 18).value = record.nov
    ws.cell(row_idx, 19).value = record.dec
    ws.cell(row_idx, 20).value = f"=SUM(P{row_idx}:S{row_idx})"
    ws.cell(row_idx, 21).value = record.jan
    ws.cell(row_idx, 22).value = record.feb
    ws.cell(row_idx, 23).value = record.mar
    ws.cell(row_idx, 24).value = record.apr
    ws.cell(row_idx, 25).value = record.may
    ws.cell(row_idx, 26).value = record.jun
    ws.cell(row_idx, 27).value = f"=SUM(U{row_idx}:Z{row_idx})"
    ws.cell(row_idx, 28).value = f"=T{row_idx}+AA{row_idx}"


def _dataframe_to_records(df: pd.DataFrame, teacher: str) -> list[PlanRecord]:
    records: list[PlanRecord] = []
    for _, row in df.iterrows():
        subject = _clean_text(row.get("subject"))
        group = _clean_text(row.get("group"))
        if not subject or not group:
            continue
        records.append(
            PlanRecord(
                subject=subject,
                group=group,
                budget=_clean_text(row.get("budget")).lower(),
                teacher=teacher,
                total_hours=_as_number_or_none(row.get("total_hours")),
                sem1_hours=_as_number_or_none(row.get("sem1_hours")),
                sem2_hours=_as_number_or_none(row.get("sem2_hours")),
                sep=_as_number_or_none(row.get("sep")),
                oct_=_as_number_or_none(row.get("oct_")),
                nov=_as_number_or_none(row.get("nov")),
                dec=_as_number_or_none(row.get("dec")),
                jan=_as_number_or_none(row.get("jan")),
                feb=_as_number_or_none(row.get("feb")),
                mar=_as_number_or_none(row.get("mar")),
                apr=_as_number_or_none(row.get("apr")),
                may=_as_number_or_none(row.get("may")),
                jun=_as_number_or_none(row.get("jun")),
            )
        )
    return records


def _write_teacher_sheet(
    ws: Worksheet,
    teacher_name: str,
    records: list[PlanRecord],
    year_start: str,
    year_end: str,
) -> None:
    data_start = 8
    row_vb = _find_label_row(ws, "Итого", column=1)
    row_budget = row_vb + 1
    row_total = _find_label_row(ws, "ВСЕГО", column=1)

    current_capacity = row_vb - data_start
    if len(records) > current_capacity:
        extra_rows = len(records) - current_capacity
        ws.insert_rows(row_vb, amount=extra_rows)
        for idx in range(row_vb, row_vb + extra_rows):
            _copy_row_style(ws, data_start, idx, 28)
        row_vb += extra_rows
        row_budget += extra_rows
        row_total += extra_rows

    ws["A3"] = f"Преподаватель  {teacher_name}  в  {year_start} - {year_end} учебном году"

    for row_idx in range(data_start, row_vb):
        for col_idx in range(1, 29):
            ws.cell(row_idx, col_idx).value = None

    for i, record in enumerate(records):
        _write_record_row(ws, data_start + i, record)

    data_end = data_start + len(records) - 1
    if data_end < data_start:
        data_end = data_start

    vb_records = [record for record in records if "вб" in record.budget]
    vb_total = _sum_values([record.total_hours for record in vb_records])
    all_total = _sum_values([record.total_hours for record in records])

    ws.cell(row_vb, 3).value = vb_total if vb_total else 0
    ws.cell(row_vb, 11).value = vb_total if vb_total else 0
    ws.cell(row_vb, 28).value = vb_total if vb_total else 0

    ws.cell(row_budget, 3).value = (all_total - vb_total) if all_total else 0
    ws.cell(row_budget, 11).value = (all_total - vb_total) if all_total else 0
    ws.cell(row_budget, 28).value = (all_total - vb_total) if all_total else 0

    for col in range(3, 29):
        letter = get_column_letter(col)
        ws.cell(row_total, col).value = f"=SUM({letter}{data_start}:{letter}{data_end})"


def build_cards_workbook(
    rup_file: str | Path,
    template_file: str | Path,
    output_file: str | Path,
    teacher_tables: dict[str, pd.DataFrame],
    selected_teachers: list[str],
) -> dict[str, int]:
    if not selected_teachers:
        raise ValueError("Не выбраны преподаватели для формирования карточек.")

    rup_path = Path(rup_file)
    template_path = Path(template_file)
    output_path = Path(output_file)

    if not rup_path.exists():
        raise FileNotFoundError(f"Файл РУП не найден: {rup_path}")
    if not template_path.exists():
        raise FileNotFoundError(f"Файл шаблона не найден: {template_path}")

    year_start, year_end = _extract_year_range(rup_path.name)
    keep_vba = template_path.suffix.lower() == ".xlsm"
    wb = load_workbook(template_path, keep_vba=keep_vba)

    template_sheet = wb["Шаблон"] if "Шаблон" in wb.sheetnames else wb[wb.sheetnames[0]]

    for sheet_name in wb.sheetnames[:]:
        if sheet_name != template_sheet.title:
            del wb[sheet_name]

    filled: dict[str, int] = {}
    existing_names: set[str] = {template_sheet.title}
    for teacher in selected_teachers:
        df = teacher_tables.get(teacher)
        if df is None:
            continue
        records = _dataframe_to_records(df, teacher)

        base_name = _sheet_name_from_teacher(teacher)
        candidate = base_name
        counter = 2
        while candidate in existing_names:
            suffix = f" {counter}"
            candidate = (base_name[: max(1, 31 - len(suffix))] + suffix).strip()
            counter += 1
        existing_names.add(candidate)

        ws = wb.copy_worksheet(template_sheet)
        ws.title = candidate
        _write_teacher_sheet(ws, teacher, records, year_start, year_end)
        filled[candidate] = len(records)

    wb.save(output_path)
    return filled
